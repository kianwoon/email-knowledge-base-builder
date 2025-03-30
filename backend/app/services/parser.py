import base64
import io
from typing import Optional

# PDF parsing
from pdfminer.high_level import extract_text

# DOCX parsing
import docx

# Excel parsing
import openpyxl


async def parse_attachment(content_bytes: str, content_type: str) -> Optional[str]:
    """
    Parse attachment content based on file type
    content_bytes: Base64 encoded content
    content_type: MIME type of the attachment
    """
    try:
        # Decode base64 content
        binary_content = base64.b64decode(content_bytes)
        
        # Parse based on content type
        if content_type == "application/pdf":
            return parse_pdf(binary_content)
        
        elif content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            return parse_docx(binary_content)
        
        elif content_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel"
        ]:
            return parse_excel(binary_content)
        
        elif content_type.startswith("text/"):
            # Plain text files
            return binary_content.decode("utf-8", errors="replace")
        
        else:
            # Unsupported file type
            return f"[Unsupported file type: {content_type}]"
    
    except Exception as e:
        print(f"Error parsing attachment: {str(e)}")
        return f"[Error parsing attachment: {str(e)}]"


def parse_pdf(content: bytes) -> str:
    """Parse PDF content to extract text"""
    try:
        # Create a file-like object from bytes
        pdf_file = io.BytesIO(content)
        
        # Extract text from PDF
        text = extract_text(pdf_file)
        return text
    
    except Exception as e:
        print(f"Error parsing PDF: {str(e)}")
        return f"[Error parsing PDF: {str(e)}]"


def parse_docx(content: bytes) -> str:
    """Parse DOCX content to extract text"""
    try:
        # Create a file-like object from bytes
        docx_file = io.BytesIO(content)
        
        # Open the document
        doc = docx.Document(docx_file)
        
        # Extract text from paragraphs
        text = "\n".join([para.text for para in doc.paragraphs])
        return text
    
    except Exception as e:
        print(f"Error parsing DOCX: {str(e)}")
        return f"[Error parsing DOCX: {str(e)}]"


def parse_excel(content: bytes) -> str:
    """Parse Excel content to extract text"""
    try:
        # Create a file-like object from bytes
        excel_file = io.BytesIO(content)
        
        # Open the workbook
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        
        # Extract text from all sheets
        text = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text.append(f"Sheet: {sheet_name}")
            
            # Extract cell values
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text.append(row_text)
            
            text.append("")  # Add a blank line between sheets
        
        return "\n".join(text)
    
    except Exception as e:
        print(f"Error parsing Excel: {str(e)}")
        return f"[Error parsing Excel: {str(e)}]"
